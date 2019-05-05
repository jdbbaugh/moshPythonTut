import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')

sheet = wb['Sheet1']
# cell = wb['a1']
cell = sheet.cell(1,1)

print(sheet.max_row)


for row in range(2, sheet.max_row + 1):
  cell = sheet.cell(row, 3)
  print(cell.value)